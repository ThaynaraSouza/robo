from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException
from openpyxl import Workbook, load_workbook
import sys
import re
import pandas as pd
import time
import os

# define pasta do arquivo
os.chdir(r'C:\Users\thayn\OneDrive\Documentos\robo_coaf')

chrome_options = webdriver.ChromeOptions()
chrome_options.add_experimental_option("useAutomationExtension", False)
chromedriver = r"C:\Users\thayn\OneDrive\Documentos\robo_coaf\chromedriver.exe"
driver = webdriver.Chrome(chromedriver, chrome_options=chrome_options)

options = Options()
options.binary_location = "C:\Program Files (x86)\Google\Chrome\Application\chrome.exe"

# df= pd.read_excel(r"C:\Users\thayn\OneDrive\Documentos\robo_coaf\extracao_coaf.xlsx",
#                 dtype={'nome_arquivo': 'str',
#                       'projeto': 'str',
#                      'link':'str'})


timeout = 3


def acesso(driver):
    try:
        driver.get('https://siscoaf.fazenda.gov.br/siscoaf-internet/pages/private/consultas/comunicacoesEnviadas.jsf')

        WebDriverWait(driver, timeout)

    except TimeoutException:
        print("Aguardando Login")
    finally:
        driver.find_element_by_id('caixa-login-certificado').click()
        time.sleep(timeout)
        print("coloque a senha")


def busca(num, comunic, driver):
    driver.get('https://siscoaf.fazenda.gov.br/siscoaf-internet/pages/private/consultas/comunicacoesEnviadas.jsf')
    driver.find_element_by_id('j_idt20:txtPessoaObrigada_label').click()

    if (comunic == 'Itaú vida e previdência S.A'):
        driver.find_element_by_xpath('//*[@id="j_idt20:txtPessoaObrigada_panel"]/div/ul/li[2]').click()
    elif (comunic == 'Itaú Seguros S.A.'):
        driver.find_element_by_xpath('//*[@id="j_idt20:txtPessoaObrigada_panel"]/div/ul/li[3]').click()
    elif (comunic == 'CIA Itaú de Capitalização'):
        driver.find_element_by_xpath('//*[@id="j_idt20:txtPessoaObrigada_panel"]/div/ul/li[8]').click()
    elif (comunic == 'Redecard S.A.'):
        driver.find_element_by_xpath('//*[@id="j_idt20:txtPessoaObrigada_panel"]/div/ul/li[11]').click()
    elif (comunic == 'ITAU ADMINISTRADORA DE CONSORCIOS LTDA'):
        driver.find_element_by_xpath('//*[@id="j_idt20:txtPessoaObrigada_panel"]/div/ul/li[18]').click()
    elif (comunic == 'Banco Itaucard S.A.'):
        driver.find_element_by_xpath('//*[@id="j_idt20:txtPessoaObrigada_panel"]/div/ul/li[19]').click()
    elif (comunic == 'Itau corretora de valores S.A.'):
        driver.find_element_by_xpath('//*[@id="j_idt20:txtPessoaObrigada_panel"]/div/ul/li[20]').click()
    elif (comunic == 'Banco Itaú S.A.'):
        driver.find_element_by_xpath('//*[@id="j_idt20:txtPessoaObrigada_panel"]/div/ul/li[26]').click()
    else:
        print('Comunicante inválido')

    driver.find_element_by_id('frmMenu:mnuSubComunicacoes').click()
    time.sleep(timeout)
    driver.find_element_by_name('txtNumeroCoaf').send_keys(num)
    time.sleep(timeout)
    driver.find_element_by_name('btnConsultar').click()
    time.sleep(timeout)

    # tbl0 = driver.find_element_by_id('listComunicacoesEnviadas_data')
    Status = driver.find_element_by_xpath('/html/body/div[2]/div[3]/div[1]/div/div/form/div[1]/div/div/div/div/div/div/table/tbody/tr/td[8]/div/span').text
    btn_pesq = driver.find_element_by_id('listComunicacoesEnviadas:0:btnVisualizarComunicacao')

    while (btn_pesq == 'None'):
        time.sleep(2)
        btn_pesq = driver.find_element_by_id('listComunicacoesEnviadas:0:btnVisualizarComunicacao')

    btn_pesq.click()
    time.sleep(2)
    
    return Status


def tb_comunicacao(num, comunic, Status, driver):

    valor = None
    valor2 = None
    valor3 = None
    valor4 = None
    valor5 = None

    segmento = driver.find_element_by_id('formViewComunicacao:j_idt209:j_idt226').text

    tbl = driver.find_element_by_id('formViewComunicacao:j_idt209:j_idt213')
    linha = driver.find_element_by_id('formViewComunicacao:j_idt209:txtOcorrencia')
    time.sleep(1)

    print(segmento)
    print(tbl)

    
    if (linha == 'None'):
        texto = driver.find_element_by_id('formViewComunicacao:j_idt209:j_idt213').text

        seg = texto.index('Segmento: ')
        tseg = len('Segmento: ')
        segmento = texto[seg + tseg + seg + tseg + 25]

        ori = texto.index('Número Origem: ')
        tori = len('Número Origem: ')
        numeroorigem = texto[ori + tori:ori + tori + 6]
        a = len(numeroorigem) - 2
        numeroorigem = numeroorigem[0:a]

        DataOperacao = driver.find_element_by_id('formViewComunicacao:j_idt209:txtDataOperacao').text
        print(DataOperacao)
        DataOperacaoFim = driver.find_element_by_id('formViewComunicacao:j_idt209:txtDataOperacaoFim').text
        print(DataOperacaoFim)
        DataRecebimento = driver.find_element_by_id('formViewComunicacao:j_idt209:txtDataRecebimento').text

        cid = texto.index('Cidade: ')
        tcidade = len('Cidade: ')
        cidade_UF = texto[cid + tcidade:cid + tcidade + 15]
    else:
        texto = driver.find_element_by_id('formViewComunicacao:j_idt209:j_idt213').text

        seg = texto.index('Segmento: ')
        tseg = len('Segmento: ')
        segmento = texto[seg + tseg + seg + tseg + 25]
        print(segmento)

        ori = texto.index('Número Origem: ')
        tori = len('Número Origem: ')
        numeroorigem = texto[ori + tori:ori + tori + 6]
        a = len(numeroorigem) - 2
        numeroorigem = numeroorigem[0:a]

        DataOperacao = driver.find_element_by_id('formViewComunicacao:j_idt209:txtDataOperacao').text
        print(DataOperacao)
        DataOperacaoFim = driver.find_element_by_id('formViewComunicacao:j_idt209:txtDataOperacaoFim').text
        print(DataOperacaoFim)
        DataRecebimento = driver.find_element_by_id('formViewComunicacao:j_idt209:txtDataRecebimento').text

        cid = texto.index('Cidade/UF: ')
        tcidade = len('Cidade/UF: ')
        cidade_UF = texto[cid + tcidade:cid + tcidade + 30]

        print(texto)
     
        valor = driver.find_element_by_id('formViewComunicacao:j_idt209:txtValor').text
        

    if (comunic == 'Banco Itaú S.A.' or comunic == 'CIA Itaú de Capitalização' or
            (comunic == 'Itaú Seguros S.A.' and Status == 'SUSEP - Mercado Segurador') or
            comunic == 'ITAU ADMINISTRADORA DE CONSORCIOS LTDA' or comunic == 'Unibanco vida e previdência S.A.' or
            comunic == 'Itau corretora de valores S.A.' and Status == 'Banco Central - Sistema Financeiro'):
        
        valor2 = driver.find_element_by_id('formViewComunicacao:j_idt209:txtValor2').text
        valor3 = driver.find_element_by_id('formViewComunicacao:j_idt209:txtValor3').text

      

    if (comunic == 'Banco Itaú S.A.' or comunic == 'ITAU ADMINISTRADORA DE CONSORCIOS LTDA' or
            (comunic == 'Banco Itaubank S.A' and Status == 'SFN - Atípicas') or
            (comunic == 'Unibanco União de Bancos Brasileiros S.A.' and Status == 'SFN - Atípicas')):

        valor4 = driver.find_element_by_id('formViewComunicacao:j_idt209:txtValor4').text
        valor5 = driver.find_element_by_id('formViewComunicacao:j_idt209:txtValor5').text

    lista_saida =pd.DataFrame([{'Num_Coaf': num, 'Numero_Origem': numeroorigem, 'Segmento': segmento,
                                   'Data_Operacao': DataOperacao, 'Data_Operacao_Fim': DataOperacaoFim,
                                   'Data_Recebimento': DataRecebimento, 'Cidade_FUF': cidade_UF,
                                   'Comunicante': comunic, 'Status': Status, 'Valor': valor, 'Valor2': valor2,
                                   'Valor3': valor3, 'Valor4': valor4, 'Valor5': valor5}])   

    return lista_saida


def tb_envolvidos(num, comunic, driver):

    objtableenvolvidos = driver.find_elements_by_xpath('//*[@id="formViewComunicacao:j_idt209:pnlConsultarComunicacao_content"]/div[2]/div[3]/table')
    qtdenvolvidos = len(objtableenvolvidos)
  
    for ii in range(0, qtdenvolvidos):

        if (driver.find_element_by_id('formViewComunicacao:j_idt209:j_idt273:' + str(ii) + ':txtCpfCnpjEnvolvido') is None ):
            tb_ocorrencias(num, driver)
        else:
            txtCpfCnpjEnvolvido = driver.find_element_by_id(
                'formViewComunicacao:j_idt209:j_idt273:' + str(ii) + ':txtCpfCnpjEnvolvido')
        if (txtCpfCnpjEnvolvido == 'None'):
            tb_ocorrencias(num, driver)

        div2 = driver.find_element_by_id('formViewComunicacao:j_idt209:pnlConsultarComunicacao_content')
        tblenv = driver.find_element_by_xpath('//*[@id="formViewComunicacao:j_idt209:gridComunicacao2"]')
        txtCpfCnpjEnvolvido = driver.find_element_by_id('formViewComunicacao:j_idt209:j_idt273:' + str(ii) + ':txtCpfCnpjEnvolvido').text
        
        txtCpfCnpjEnvolvido = re.sub('[^0-9]', '', txtCpfCnpjEnvolvido)
        

        if len(txtCpfCnpjEnvolvido) > 11:
            txtCpfCnpjEnvolvido.format('00000000000000')
        else:
            txtCpfCnpjEnvolvido.format('00000000000')
        
        txtNomeEnvolvido = driver.find_element_by_id('formViewComunicacao:j_idt209:j_idt273:' + str(ii) +':txtNomeEnvolvido').text
        
        txtTipoEnvolvimento = driver.find_element_by_id('formViewComunicacao:j_idt209:j_idt273:' + str(ii) +':txtTipoEnvolvimento').text
        
        txtPEP = driver.find_element_by_id('formViewComunicacao:j_idt209:j_idt273:' + str(ii) +':txtEnvolvidoPEP').text
        
        txtPO = driver.find_element_by_id('formViewComunicacao:j_idt209:j_idt273:' + str(ii) +':txtEnvolvidoPessoaObrigada').text
        
        txtSP = driver.find_element_by_id('formViewComunicacao:j_idt209:j_idt273:' + str(ii) +':txtEnvolvidoPessoaServidorPublico').text
        
        lista_saida = pd.DataFrame([{'CpfCnpjEnvolvido': txtCpfCnpjEnvolvido, 'NomeEnvolvido': txtNomeEnvolvido,
                                    'TipoEnvolvimento': txtTipoEnvolvimento, 'PEP': txtPEP,
                                    'PessoaObrigada': txtPO, 'ServidorPublico': txtSP, 
                                    'NumeroCoaf': num, 'Comunicante': comunic}])
    #print(lista_saida)
    return lista_saida


def tb_ocorrencias(num, comunic, driver):
  
    
    objtablealinea = driver.find_elements_by_xpath('//*[@id="formViewComunicacao:j_idt209:pnlConsultarComunicacao_content"]/div[2]/div[1]')
    qtdalinea = len(objtablealinea)
    for cont1 in range(0, qtdalinea):
        #qtdelinha = qtdelinha + 1
        #if (('formViewComunicacao:j_idt209:j_idt265:' + str(cont1) + ':txtOcorrencia') != 'None'):

        if ((driver.find_element_by_id('formViewComunicacao:j_idt209:j_idt265:' + str(cont1) + ':txtOcorrencia')) == 'None'):
                tb_envolvidos(num, driver)
        else:

            txtOcorrencia = driver.find_element_by_id('formViewComunicacao:j_idt209:j_idt265:' + str(cont1) + ':txtOcorrencia')
            print(txtOcorrencia)

        if (txtOcorrencia == 'None'):
            tb_envolvidos(num, driver)
            
        txtOcorrencia = driver.find_element_by_id('formViewComunicacao:j_idt209:j_idt265:' + str(cont1) + ':txtOcorrencia').text

        aux = txtOcorrencia[0:3]
        numOcorrencia = aux

        lista_saida = pd.DataFrame([{'TextoAlinea': txtOcorrencia, 'NumAlinea': numOcorrencia,
                                    'NumeroCoaf': num, 'Comunicante': comunic}])

    driver.find_element_by_tag_name('body').send_keys(Keys.CONTROL + 't')
            # driver.get('https://siscoaf.fazenda.gov.br/siscoaf-internet/pages/private/consultas/comunicacoesEnviadas.jsf')

    print(lista_saida) 
    return lista_saida

    # planilha4 = arquivo_excel.create_sheet("Informações Adicionais")
    # planilha4['A2'] = 'num'




"""
#def gera_tb():
 #   global num
  #  global comunic
   # global qtdelinha

    #wb = load_workbook('Comunicação.xlsx')
    #source = wb['Lista']
    
    qtdelinha = 1
    for i in source['A']:
        num = (i.value)
        for j in source['B']:
            comunic = (j.value)
            qtdelinha = qtdelinha + 1
            # num = '22772174'
            # comunic = 'Itaú vida e previdência S.A'

            busca(num, driver)
            tb_comunicacao(num, driver)
            tb_envolvidos(num, driver)
            tb_ocorrencias(num, driver)

    # arquivo_excel.save("Extração_Coaf.xlsx")
            
"""
acesso(driver)
#gera_tb()

##############################################################################


# lê o arquivo de entrada
entrada = pd.read_excel(open('Comunicacoes.xlsx', 'rb'), sheet_name='Lista') 

# inicialza o arquivo para saída
saida_comunic = pd.DataFrame()
saida_env = pd.DataFrame()
saida_alinea = pd.DataFrame()

# faz o loop para acessar o site para cada linha de entrada
for i,n_comunicacao in enumerate(entrada.Comunicacoes):
    print(n_comunicacao)
    
    Status = busca(n_comunicacao, entrada.iloc[i, 1], driver)

    # Ponha seu código de web scraping

    # variáveis usadas no web scraping
    tb_comunic = tb_comunicacao(n_comunicacao, entrada.iloc[i, 1], Status, driver)
    tb_env = tb_envolvidos(n_comunicacao, entrada.iloc[i, 1], driver)
    tb_alinea = tb_ocorrencias(n_comunicacao, entrada.iloc[i, 1], driver )

    # appends no arquivo para saida
    saida_comunic = saida_comunic.append(tb_comunic, ignore_index=True)
    saida_env = saida_env.append(tb_env, ignore_index=True)
    saida_alinea = saida_alinea.append(tb_alinea, ignore_index=True)

#left join
     #pd.merge(tb_comunic, tb_env, how='left', on=[num])
# Salva saida
saida_comunic.to_excel('Saida_Comunicacoes.xlsx', sheet_name = 'Comunicados', index = False)
saida_env.to_excel('Envolvidos.xlsx', sheet_name = 'Envolvidos', index = False)
saida_alinea.to_excel('Alíneas.xlsx', sheet_name = 'Ocorrências', index = False)

